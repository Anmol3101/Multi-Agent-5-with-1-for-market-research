"""
Dynamic Excel builder — works with any final_json from HYPERMARKET 5+1.
Builds sheets from structured data regardless of product or market.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY   = "1A2B4A"
ORANGE = "E87722"
WHITE  = "FFFFFF"
LG     = "F2F2F2"
RED    = "FFCCCC"
YLW    = "FFFACC"
GRN    = "CCFFCC"
BLUE   = "CCE5FF"

hf   = Font(name="Arial", bold=True, color=WHITE, size=10)
bf   = Font(name="Arial", size=10)
bdf  = Font(name="Arial", bold=True, size=10)
CA   = Alignment(horizontal="center", vertical="center", wrap_text=True)
LA   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
thin = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin")
)


def _H(cell):
    cell.font = hf
    cell.fill = PatternFill("solid", start_color=NAVY)
    cell.alignment = CA
    cell.border = thin


def _B(cell, bold=False, fc=None, a=None):
    cell.font = bdf if bold else bf
    cell.alignment = a or CA
    cell.border = thin
    if fc:
        cell.fill = PatternFill("solid", start_color=fc)


def _W(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_sheet(ws, headers, rows, col_widths, col_alignments=None):
    """Write a generic sheet with headers + rows."""
    ws.row_dimensions[1].height = 28
    for c, h in enumerate(headers, 1):
        _H(ws.cell(1, c, h))
    for r, row in enumerate(rows, 2):
        fc = LG if r % 2 == 0 else None
        for c, v in enumerate(row, 1):
            a = (col_alignments[c - 1] if col_alignments else CA)
            _B(ws.cell(r, c, str(v) if v is not None else ""), fc=fc, a=a)
    _W(ws, col_widths)


# ── Sheet builders ────────────────────────────────────────────────────────────

def _sheet_summary(wb, fj):
    ws = wb.active
    ws.title = "Summary Dashboard"
    mo = fj.get("market_overview", {})
    an = fj.get("analysis", {})
    rb = mo.get("regional_breakdown", {})
    pr = an.get("pricing_recommendation", {})

    markets = list(rb.keys()) if rb else []
    headers = ["Metric"] + [m.upper() for m in markets] if markets else ["Metric", "Value"]

    rows = []

    size = mo.get("estimated_market_size", {})
    if size:
        rows.append(["Market Size Est. ({} {})".format(
            size.get("currency", "USD"), size.get("unit", "M")
        )] + [rb.get(m, {}).get("market_size_usd_million", "—") for m in markets])

    growth = mo.get("estimated_growth_rate", {})
    if growth:
        rows.append(["CAGR % ({})".format(growth.get("period", ""))] +
                    [rb.get(m, {}).get("cagr_percent", "—") for m in markets])

    if rb:
        rows.append(["Maturity Stage"] +
                    [rb.get(m, {}).get("maturity", "—").title() for m in markets])

    gtm = an.get("gtm_recommendations", {})
    if gtm:
        rows.append(["GTM Entry Priority"] +
                    [gtm.get(m, {}).get("priority", "—") for m in markets])

    if pr:
        rows.append(["Rec. Price Range"] +
                    ["{} {}–{}".format(
                        pr.get(m, {}).get("currency", ""),
                        pr.get(m, {}).get("min", ""),
                        pr.get(m, {}).get("max", "")
                    ) for m in markets])
        rows.append(["Price Sweet Spot"] +
                    [pr.get(m, {}).get("sweet_spot", "—") for m in markets])

    ds = an.get("overall_demand_score", {})
    if ds:
        rows.append(["Demand Score (1–10)"] + [ds.get("score_1_to_10", "—")] * len(markets))

    if gtm:
        rows.append(["Top Launch Channel"] +
                    [", ".join(gtm.get(m, {}).get("channels", [])[:2]) for m in markets])

    n_cols = len(headers)
    _write_sheet(ws, headers, rows,
                 [30] + [22] * (n_cols - 1),
                 [LA] + [CA] * (n_cols - 1))


def _sheet_competitors(wb, fj):
    ws = wb.create_sheet("Competitor Matrix")
    comps = fj.get("competitors", [])
    if not comps:
        ws["A1"] = "No competitor data available."
        return
    headers = ["Brand", "Type", "Markets", "Price Band", "Positioning", "Primary Channels", "Website"]
    rows = [
        [
            c.get("name", ""),
            c.get("type", ""),
            ", ".join(c.get("markets", [])) if isinstance(c.get("markets"), list) else c.get("markets", ""),
            c.get("price_band", ""),
            c.get("positioning", ""),
            ", ".join(c.get("channels", [])),
            c.get("website", ""),
        ]
        for c in comps
    ]
    _write_sheet(ws, headers, rows,
                 [24, 12, 16, 20, 38, 36, 20],
                 [LA, CA, CA, CA, LA, LA, CA])


def _sheet_segments(wb, fj):
    ws = wb.create_sheet("Segment Opportunities")
    segs = fj.get("analysis", {}).get("segment_opportunities", [])
    if not segs:
        ws["A1"] = "No segment data available."
        return
    headers = ["Segment Name", "Target Market", "Core Pain Point", "Suggested Positioning", "WTP Score (1–10)"]
    rows = [
        [
            s.get("segment_name", ""),
            s.get("target_market", ""),
            s.get("need_or_pain_point", ""),
            s.get("suggested_positioning", ""),
            s.get("willingness_to_pay_1_to_10", ""),
        ]
        for s in segs
    ]
    _write_sheet(ws, headers, rows,
                 [28, 18, 40, 40, 16],
                 [LA, CA, LA, LA, CA])


def _sheet_risks(wb, fj):
    ws = wb.create_sheet("Risk Register")
    risks = fj.get("analysis", {}).get("key_risks_summary", [])
    if not risks:
        ws["A1"] = "No risk data available."
        return
    level_colors = {"high": RED, "medium": YLW, "low": GRN}
    headers = ["Risk Description", "Risk Level", "Markets Affected", "Notes"]
    ws.row_dimensions[1].height = 28
    for c, h in enumerate(headers, 1):
        _H(ws.cell(1, c, h))
    for r, risk in enumerate(risks, 2):
        level = str(risk.get("level", "")).lower()
        fc_row = LG if r % 2 == 0 else None
        vals = [
            risk.get("description", ""),
            risk.get("level", "").title(),
            ", ".join(risk.get("markets", [])) if isinstance(risk.get("markets"), list) else risk.get("markets", ""),
            "",
        ]
        for c, v in enumerate(vals, 1):
            fc = level_colors.get(level) if c == 2 else fc_row
            _B(ws.cell(r, c, str(v)), fc=fc, a=CA if c == 2 else LA)
    _W(ws, [44, 12, 22, 30])


def _sheet_gtm(wb, fj):
    ws = wb.create_sheet("GTM Roadmap")
    gtm = fj.get("analysis", {}).get("gtm_recommendations", {})
    if not gtm:
        ws["A1"] = "No GTM data available."
        return
    market_colors = {0: "CCE5CC", 1: "CCE0FF", 2: "FFE5CC", 3: "FFE8E8", 4: "E8E8FF"}
    headers = ["Market", "Priority", "Channels", "Core Strategy", "Timeline"]
    rows = []
    for i, (market, data) in enumerate(sorted(gtm.items(), key=lambda x: x[1].get("priority", 99))):
        rows.append([
            market.upper(),
            data.get("priority", ""),
            ", ".join(data.get("channels", [])[:3]),
            data.get("strategy", ""),
            data.get("timeline", ""),
        ])
    ws.row_dimensions[1].height = 28
    for c, h in enumerate(headers, 1):
        _H(ws.cell(1, c, h))
    for r, (row, (market, _)) in enumerate(zip(rows, sorted(gtm.items(), key=lambda x: x[1].get("priority", 99))), 2):
        idx = list(sorted(gtm.keys(), key=lambda k: gtm[k].get("priority", 99))).index(market)
        fc = market_colors.get(idx % 5, LG)
        for c, v in enumerate(row, 1):
            _B(ws.cell(r, c, str(v)), bold=(c in (1, 2)), fc=fc,
               a=CA if c in (1, 2) else LA)
    _W(ws, [12, 10, 32, 50, 28])


def _sheet_trends(wb, fj):
    trends = fj.get("analysis", {}).get("key_trends", [])
    if not trends:
        return
    ws = wb.create_sheet("Market Trends")
    relevance_colors = {"high": BLUE, "medium": YLW, "low": GRN}
    headers = ["Trend", "Relevance", "Markets Affected"]
    ws.row_dimensions[1].height = 28
    for c, h in enumerate(headers, 1):
        _H(ws.cell(1, c, h))
    for r, t in enumerate(trends, 2):
        rel = str(t.get("relevance", "")).lower()
        fg = LG if r % 2 == 0 else None
        vals = [
            t.get("trend", ""),
            t.get("relevance", "").title(),
            ", ".join(t.get("markets", [])) if isinstance(t.get("markets"), list) else t.get("markets", ""),
        ]
        for c, v in enumerate(vals, 1):
            fc = relevance_colors.get(rel) if c == 2 else fg
            _B(ws.cell(r, c, str(v)), fc=fc, a=CA if c == 2 else LA)
    _W(ws, [58, 12, 26])


def _sheet_meta(wb, fj):
    ws = wb.create_sheet("Meta & Open Questions")
    meta = fj.get("meta", {})
    oq = meta.get("open_questions", [])
    conf = meta.get("confidence", {})

    headers = ["Item", "Value"]
    rows = [
        ["Product",       fj.get("product", "")],
        ["Target Market", fj.get("target_market", "")],
        ["Timestamp UTC", fj.get("timestamp_utc", "")],
        ["Data Quality",  conf.get("data_quality", "").title()],
        ["Numeric Accuracy", conf.get("numeric_accuracy", "").title()],
        ["Strategic Assessment", conf.get("strategic_assessment", "").title()],
        ["Notes",         meta.get("notes", "")],
    ]
    if oq:
        rows.append(["", ""])
        rows.append(["OPEN QUESTIONS", ""])
        for q in oq:
            rows.append(["", q])

    ws.row_dimensions[1].height = 28
    for c, h in enumerate(headers, 1):
        _H(ws.cell(1, c, h))
    for r, row in enumerate(rows, 2):
        fc = LG if r % 2 == 0 else None
        for c, v in enumerate(row, 1):
            bold = (row[0] == "OPEN QUESTIONS" and c == 1)
            _B(ws.cell(r, c, str(v)), bold=bold, fc=fc, a=LA)
    _W(ws, [28, 70])


# ── Public API ────────────────────────────────────────────────────────────────

def build_excel(final_json: dict, output_path: str) -> None:
    """Build a full Excel workbook from any HYPERMARKET final_json output."""
    wb = Workbook()
    _sheet_summary(wb, final_json)
    _sheet_competitors(wb, final_json)
    _sheet_segments(wb, final_json)
    _sheet_risks(wb, final_json)
    _sheet_gtm(wb, final_json)
    _sheet_trends(wb, final_json)
    _sheet_meta(wb, final_json)
    wb.save(output_path)
